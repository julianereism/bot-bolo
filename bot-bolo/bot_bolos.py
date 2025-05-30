import os
import logging
from telegram import Update, ReplyKeyboardMarkup
from telegram.ext import (
    ApplicationBuilder, CommandHandler, MessageHandler,
    filters, ConversationHandler, ContextTypes
)
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook

# Carrega o token do .env
load_dotenv()
TOKEN = os.getenv("TELEGRAM_API")

# Estados da conversa
MENU, ESCOLHER_BOLO, INFORMAR_NOME, INFORMAR_ENDERECO = range(4)

# Opções de bolos
bolos = ["Bolo de Chocolate", "Bolo de Morango", "Bolo de Cenoura", "Bolo de Leite Ninho"]

# Configura o logger
logging.basicConfig(level=logging.INFO)

# Função para salvar pedido na planilha Excel
def salvar_pedido_excel(nome, bolo, endereco):
    caminho = "pedidos.xlsx"
    if not os.path.exists(caminho):
        wb = Workbook()
        ws = wb.active
        ws.append(["Nome", "Bolo", "Endereço", "Status"])
        wb.save(caminho)

    wb = load_workbook(caminho)
    ws = wb.active
    ws.append([nome, bolo, endereco, "Aguardando pagamento"])
    wb.save(caminho)

# Teclado de voltar
def teclado_voltar():
    return [["Voltar ao menu"]]

# Início do bot
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    nome = update.effective_user.first_name
    mensagem = f"Olá, {nome}! 🎂 Bem-vindo ao *Delivery de Bolos da Nana Reis*.\n\nEscolha uma opção:"
    teclado = [["1. Fazer pedido", "2. Contato e Pagamento"]]
    await update.message.reply_text(mensagem, parse_mode="Markdown", reply_markup=ReplyKeyboardMarkup(teclado, resize_keyboard=True))
    return MENU

# Menu principal
async def menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    texto = update.message.text.lower()

    if "1" in texto:
        teclado = [[b] for b in bolos] + teclado_voltar()
        await update.message.reply_text("Escolha o sabor do bolo:", reply_markup=ReplyKeyboardMarkup(teclado, resize_keyboard=True))
        return ESCOLHER_BOLO

    elif "2" in texto:
        mensagem = (
            "📞 *Contato:*\nWhatsApp: (81) 99959-7056\nInstagram: @d__nanareis\n\n"
            "💳 *Pagamento via PIX:*\n"
            "Chave: `pagamento@nanareis.com`\n"
            "_Envie o comprovante pelo WhatsApp após finalizar o pedido._"
        )
        await update.message.reply_text(mensagem, parse_mode="Markdown", reply_markup=ReplyKeyboardMarkup(teclado_voltar(), resize_keyboard=True))
        return MENU

    elif "voltar" in texto:
        return await start(update, context)
    else:
        await update.message.reply_text("Não entendi. Escolha uma opção do menu.")
        return MENU

# Escolha do bolo
async def escolher_bolo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    bolo = update.message.text
    if bolo.lower() == "voltar":
        return await start(update, context)
    if bolo not in bolos:
        await update.message.reply_text("Por favor, escolha um sabor do teclado.")
        return ESCOLHER_BOLO

    context.user_data["bolo"] = bolo
    await update.message.reply_text("Digite seu *nome completo*:", parse_mode="Markdown", reply_markup=ReplyKeyboardMarkup(teclado_voltar(), resize_keyboard=True))
    return INFORMAR_NOME

# Nome
async def informar_nome(update: Update, context: ContextTypes.DEFAULT_TYPE):
    nome = update.message.text
    if nome.lower() == "voltar":
        return await start(update, context)

    context.user_data["nome"] = nome
    await update.message.reply_text("Informe o *endereço de entrega* (com rua, número e bairro):", parse_mode="Markdown", reply_markup=ReplyKeyboardMarkup(teclado_voltar(), resize_keyboard=True))
    return INFORMAR_ENDERECO

# Endereço
async def informar_endereco(update: Update, context: ContextTypes.DEFAULT_TYPE):
    endereco = update.message.text
    if endereco.lower() == "voltar":
        return await start(update, context)

    context.user_data["endereco"] = endereco
    return await confirmar_pedido(update, context)

# Confirmação e pagamento
async def confirmar_pedido(update: Update, context: ContextTypes.DEFAULT_TYPE):
    nome = context.user_data["nome"]
    bolo = context.user_data["bolo"]
    endereco = context.user_data["endereco"]

    salvar_pedido_excel(nome, bolo, endereco)

    chave_pix = "pagamento@nanareis.com"  # Troque pela sua chave PIX

    mensagem = (
        f"✅ *Pedido registrado!*\n\n"
        f"🍰 Bolo: *{bolo}*\n"
        f"👤 Nome: *{nome}*\n"
        f"📍 Endereço: *{endereco}*\n\n"
        "💳 *Pagamento via PIX:*\n"
        f"🔑 Chave: `{chave_pix}`\n"
        "_Envie o comprovante via WhatsApp para confirmar o pedido._\n\n"
        "📞 WhatsApp: (81) 99959-7056"
    )
    await update.message.reply_text(mensagem, parse_mode="Markdown", reply_markup=ReplyKeyboardMarkup(teclado_voltar(), resize_keyboard=True))
    return MENU

# Fallback
async def fallback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Desculpe, não entendi. Escolha uma opção do menu.", reply_markup=ReplyKeyboardMarkup(teclado_voltar(), resize_keyboard=True))
    return MENU

# Inicialização do bot
def main():
    app = ApplicationBuilder().token(TOKEN).build()
    app.add_handler(CommandHandler("start", start))

    conv_handler = ConversationHandler(
        entry_points=[MessageHandler(filters.TEXT & ~filters.COMMAND, menu)],
        states={
            MENU: [MessageHandler(filters.TEXT & ~filters.COMMAND, menu)],
            ESCOLHER_BOLO: [MessageHandler(filters.TEXT & ~filters.COMMAND, escolher_bolo)],
            INFORMAR_NOME: [MessageHandler(filters.TEXT & ~filters.COMMAND, informar_nome)],
            INFORMAR_ENDERECO: [MessageHandler(filters.TEXT & ~filters.COMMAND, informar_endereco)],
        },
        fallbacks=[MessageHandler(filters.TEXT, fallback)],
    )

    app.add_handler(conv_handler)
    print("🎂 Bot de delivery de bolos rodando...")
    app.run_polling()

if __name__ == "__main__":
    main()
